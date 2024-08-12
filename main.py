from fastapi import FastAPI
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from fastapi.middleware.cors import CORSMiddleware
import os
import math
import re
app = FastAPI()

# CORS middleware configuration
origins = [
    "http://localhost:3000",  # 允许的来源
    "http://127.0.0.1:3000",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,  # 允许的来源列表
    allow_credentials=True,
    allow_methods=["*"],  # 允许的 HTTP 方法列表
    allow_headers=["*"],  # 允许的 HTTP 头列表
)

def convert_to_float(value):
    try:
        return float(value)
    except ValueError:
        return math.nan  # 将无法转换的值设置为 NaN

def clean_data_for_json(data):
    # 清除无效的 JSON 值
    for item in data:
        if math.isnan(item['importance']):
            item['importance'] = None
    return data

@app.get("/readfile/")
async def read_excel_file():
    # Define the path to the Excel file
    file_path = os.path.join("excel", "importances.xlsx")
    
    # Load the workbook and select the active sheet
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook.active

    # Extract data from the sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        if row[0] and row[1]:  # Skip rows with empty feature or importance
            data.append({
                "feature": re.sub('_resampled','',row[0]),
                "importance": convert_to_float(row[1])  # Convert importance to float
            })

    # Sort data by importance, putting None values at the end
    data.sort(key=lambda x: x['importance'] if not math.isnan(x['importance']) else float('inf'), reverse=True)

    # Clean the data for JSON response
    cleaned_data = clean_data_for_json(data)

    return JSONResponse(content={"data": cleaned_data})

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
